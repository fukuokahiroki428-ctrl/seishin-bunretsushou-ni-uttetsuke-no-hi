#!/usr/bin/env python3
"""Bluesky collection daemon for ABIWA - based on AINU (ainu_smart.py).

Usage: python3 bluesky_daemon.py '{"handle":"...","password":"..."}'
Then send JSON commands on stdin, receive JSON responses on stdout.

Commands:
  {"action":"posts","target":"handle","save_path":"/path","download_media":true,"exif":true}
  {"action":"likes","target":"handle","save_path":"/path","download_media":true}
  {"action":"media","target":"handle","save_path":"/path"}
  {"action":"followers","target":"handle","save_path":"/path"}
  {"action":"following","target":"handle","save_path":"/path"}
  {"action":"blocks","save_path":"/path"}
  {"action":"mutes","save_path":"/path"}
  {"action":"lists","target":"handle","save_path":"/path"}
  {"action":"notifications","save_path":"/path"}
  {"action":"messages","save_path":"/path"}
  {"action":"search","query":"keyword","save_path":"/path","download_media":true}
  {"action":"profile","target":"handle","save_path":"/path"}
  {"action":"ping"}
  {"action":"quit"}
"""

import sys, json, os, time, asyncio, re, subprocess, shutil
from datetime import datetime, timezone
from pathlib import Path

def log(msg):
    print(json.dumps({"log": str(msg)}), flush=True)

def progress(count, media_count=0, status="수집 중..."):
    print(json.dumps({"progress": {"count": count, "media": media_count, "status": status}}), flush=True)

def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Usage: bluesky_daemon.py '{\"handle\":\"...\",\"password\":\"...\"}'"}), flush=True)
        sys.exit(1)

    init_args = json.loads(sys.argv[1])

    # Support multiple accounts: {"accounts":[{"handle":"...","password":"..."},...]}
    # or single account: {"handle":"...","password":"..."}
    accounts = init_args.get("accounts", [])
    if not accounts:
        accounts = [{"handle": init_args["handle"], "password": init_args["password"]}]

    # Import dependencies
    try:
        from atproto import Client as ATProtoClient
    except ImportError:
        print(json.dumps({"error": "atproto not installed. pip install atproto"}), flush=True)
        sys.exit(1)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print(json.dumps({"error": "openpyxl not installed"}), flush=True)
        sys.exit(1)

    # Login with first account
    current_account_idx = 0
    clients = []
    for i, acct in enumerate(accounts):
        try:
            c = ATProtoClient()
            c.login(acct["handle"], acct["password"])
            clients.append(c)
            log(f"Account {i+1} logged in: {acct['handle']}")
        except Exception as e:
            log(f"Account {i+1} login failed: {acct['handle']} - {e}")

    if not clients:
        print(json.dumps({"error": f"Login failed: all accounts failed"}), flush=True)
        sys.exit(1)

    client = clients[0]

    def rotate_account():
        """Rate limit 시 다음 계정으로 순환"""
        nonlocal client, current_account_idx
        if len(clients) <= 1:
            return False
        current_account_idx = (current_account_idx + 1) % len(clients)
        client = clients[current_account_idx]
        log(f"🔄 계정 전환: {accounts[current_account_idx]['handle']} ({current_account_idx+1}/{len(clients)})")
        return True

    rate_limit_count = 0
    total_rate_limit_waits = 0
    MAX_TOTAL_WAITS = 20  # 대기를 최대 20회까지 허용
    adaptive_delay = 1.0  # 아이누 방식: 기본 1초 고정

    def get_delay():
        """아이누 방식: 고정 딜레이 (기본 1초, RL 후 일시 증가)"""
        return adaptive_delay

    def decay_delay():
        """아이누 방식: RL 후 증가한 딜레이를 원래대로"""
        nonlocal adaptive_delay
        adaptive_delay = 1.0  # 아이누: 항상 1초로 리셋

    # Flag for graceful shutdown during rate limit waits
    stop_requested = False

    def check_stdin_for_quit():
        """Non-blocking stdin check for quit command during rate limit waits."""
        nonlocal stop_requested
        import select
        while select.select([sys.stdin], [], [], 0)[0]:
            try:
                line = sys.stdin.readline().strip()
                if not line:
                    continue
                cmd = json.loads(line)
                if cmd.get("action") == "quit":
                    stop_requested = True
                    return True
            except:
                pass
        return False

    # Rate limit mode: "wait" = 대기 후 재시도, "stop" = 즉시 중지
    rl_mode = "wait"  # default
    rl_wait_secs = 900  # default 15분

    def set_rate_limit_mode(mode, wait_mins=15):
        nonlocal rl_mode, rl_wait_secs
        rl_mode = mode  # "wait" or "stop"
        rl_wait_secs = max(60, min(3600, wait_mins * 60))

    def handle_rate_limit():
        """AINU 방식 Rate limit: 대기 → 중간 저장 → 다음 계정 전환"""
        nonlocal rate_limit_count, total_rate_limit_waits, client, adaptive_delay, stop_requested
        rate_limit_count += 1

        # "stop" 모드면 즉시 중지
        if rl_mode == "stop":
            log("⚠️ Rate Limit 발생 → 수집 중지 (설정: 중지 모드)")
            return False

        wait_secs = rl_wait_secs
        wait_mins = wait_secs // 60
        log(f"⚠️ Rate Limit 발생")
        log(f"⏰ {wait_mins}분 대기... (중지하려면 앱에서 중지 버튼을 누르세요)")

        # 대기 — 1초마다 stdin 체크해서 중지 가능하게
        for i in range(wait_secs):
            if stop_requested or check_stdin_for_quit():
                log("⏹️ 사용자에 의해 중지됨")
                return False
            remaining = wait_secs - i
            if remaining % 10 == 0:
                mins = remaining // 60
                secs = remaining % 60
                progress(0, 0, f"Rate Limit 대기... {mins}분 {secs}초")
            time.sleep(1)

        adaptive_delay = 1.0  # 대기 완료 → 딜레이 리셋
        log("✅ 대기 완료! 다음 계정으로 전환...")

        # 다음 계정으로 전환 (AINU 방식)
        if rotate_account():
            # 재로그인
            try:
                acct = accounts[current_account_idx]
                clients[current_account_idx] = ATProtoClient()
                clients[current_account_idx].login(acct["handle"], acct["password"])
                client = clients[current_account_idx]
                log(f"🔄 계정 전환: {acct['handle']}")
            except Exception as e:
                log(f"⚠️ 계정 전환 실패: {e}")
            return True
        else:
            # 계정 1개: 15분 대기 후 재로그인 시도
            total_rate_limit_waits += 1
            if total_rate_limit_waits > MAX_TOTAL_WAITS:
                log("❌ Rate Limit 소진. 나중에 다시 시도해주세요.")
                return False
            try:
                acct = accounts[current_account_idx]
                clients[current_account_idx] = ATProtoClient()
                clients[current_account_idx].login(acct["handle"], acct["password"])
                client = clients[current_account_idx]
                log("✅ 재로그인 성공! 수집 재개...")
            except Exception as e:
                log(f"⚠️ 재로그인 실패: {e}, 그래도 계속 시도...")
            return True

    print(json.dumps({"status": "ready"}), flush=True)

    # Helper: normalize handle
    def norm_handle(h):
        h = h.strip().lstrip('@')
        short = h.replace('.bsky.social', '')
        api = h if '.' in h else h + '.bsky.social'
        return short, api

    # Helper: process a post into a data row
    def process_post(post):
        """Extract data from a Bluesky post. Returns list of 11 fields."""
        try:
            uri = getattr(post, 'uri', '') or ''
            cid = getattr(post, 'cid', '') or ''
            author = getattr(post, 'author', None)
            record = getattr(post, 'record', None)

            author_handle = getattr(author, 'handle', '') if author else ''
            author_name = getattr(author, 'display_name', '') if author else ''

            text = getattr(record, 'text', '') if record else ''
            text = text.replace('\n', ' ').replace('\r', '')
            lang_list = getattr(record, 'langs', []) if record else []
            language = lang_list[0] if lang_list else ''

            created_str = getattr(record, 'created_at', '') if record else ''
            # Convert to JST
            created_at = ''
            if created_str:
                try:
                    dt = datetime.fromisoformat(created_str.replace('Z', '+00:00'))
                    from datetime import timedelta
                    jst = dt.astimezone(timezone(timedelta(hours=9)))
                    created_at = jst.strftime('%Y/%m/%d %H:%M:%S')
                except:
                    created_at = str(created_str)

            like_count = getattr(post, 'like_count', 0) or 0
            repost_count = getattr(post, 'repost_count', 0) or 0
            reply_count = getattr(post, 'reply_count', 0) or 0

            # Media
            embed = getattr(post, 'embed', None)
            has_media = False
            media_count = 0
            media_urls = []

            if embed:
                images = getattr(embed, 'images', None)
                if images:
                    has_media = True
                    for img in images:
                        url = getattr(img, 'fullsize', '') or getattr(img, 'thumb', '')
                        if url:
                            media_urls.append(url)
                            media_count += 1

                video = getattr(embed, 'video', None)
                if video:
                    has_media = True
                    playlist = getattr(video, 'playlist', '')
                    if playlist:
                        media_urls.append(playlist)
                        media_count += 1

                # Record with media (quote + images)
                rmedia = getattr(embed, 'media', None)
                if rmedia:
                    rimages = getattr(rmedia, 'images', None)
                    if rimages:
                        has_media = True
                        for img in rimages:
                            url = getattr(img, 'fullsize', '') or getattr(img, 'thumb', '')
                            if url:
                                media_urls.append(url)
                                media_count += 1

                ext = getattr(embed, 'external', None)
                if ext:
                    thumb = getattr(ext, 'thumb', '')
                    if thumb:
                        has_media = True
                        media_urls.append(thumb)
                        media_count += 1

            # Post URL
            post_id = uri.split('/')[-1] if uri else ''
            post_url = f"https://bsky.app/profile/{author_handle}/post/{post_id}" if author_handle and post_id else ''

            return [uri, post_url, text, language, created_at,
                    author_handle, author_name,
                    like_count, repost_count, reply_count,
                    'Yes' if has_media else 'No', media_count,
                    '\n'.join(media_urls)]
        except Exception as e:
            return None

    # Helper: download media files
    def download_media_files(post, media_dir, post_data, action_type="posts"):
        """Download media from a post. Returns count of new files.
        action_type: posts / likes / search — per-type subfolder + _complete mirror."""
        import urllib.request, shutil
        embed = getattr(post, 'embed', None)
        if not embed:
            return 0

        uri = getattr(post, 'uri', '') or ''
        post_id = uri.split('/')[-1] if uri else 'unknown'
        count = 0

        # Author username subfolder
        author = getattr(post, 'author', None)
        author_handle = getattr(author, 'handle', '') if author else ''
        if author_handle:
            short_author = author_handle.replace('.bsky.social', '')
            media_dir = os.path.join(media_dir, short_author)

        # Per-type subfolder (posts/, likes/, search/)
        type_dir = os.path.join(media_dir, action_type)
        os.makedirs(type_dir, exist_ok=True)
        # _complete mirror folder
        complete_dir = os.path.join(media_dir, "_complete")
        os.makedirs(complete_dir, exist_ok=True)
        # Override media_dir to type subfolder
        media_dir = type_dir

        # 파일명 생성: 게시물텍스트 (post_id_날짜).확장자
        def make_filename(suffix, ext):
            import re
            text = ''
            if post_data and len(post_data) > 1:
                text = str(post_data[1])[:40].strip().replace('\n', ' ')
            text = re.sub(r'[/\\:*?"<>|\x00-\x1f]', '_', text)
            if not text:
                text = post_id
            date_str = ''
            if post_data and len(post_data) > 4 and post_data[4]:
                try:
                    dt = datetime.strptime(post_data[4], '%Y/%m/%d %H:%M:%S')
                    date_str = dt.strftime('%Y%m%d_%H%M')
                except:
                    pass
            # 업로드 시각 prefix → OS 기본 정렬로 업로드 순 배치
            order_prefix = f"{date_str}_" if date_str else ""
            name = f"{order_prefix}{text} ({post_id}"
            if suffix:
                name += f"_{suffix}"
            name += f"){ext}"
            return name[:200]

        def dl(url, filename):
            nonlocal count
            filepath = os.path.join(media_dir, filename)
            if os.path.exists(filepath):
                return filepath
            try:
                urllib.request.urlretrieve(url, filepath)
                count += 1
                # Set file date
                if post_data and len(post_data) > 4 and post_data[4]:
                    try:
                        from datetime import timedelta
                        dt = datetime.strptime(post_data[4], '%Y/%m/%d %H:%M:%S')
                        dt = dt.replace(tzinfo=timezone(timedelta(hours=9)))
                        ts = dt.timestamp()
                        os.utime(filepath, (ts, ts))
                        # macOS creation date
                        mac_date = dt.strftime('%m/%d/%Y %H:%M:%S')
                        # ★ os.system + f-string 금지 — filepath 에 " $ ` \ 가 있으면 셸 주입.
                        #   argv 로 직접 실행(셸 미경유). SetFile 은 legacy 라 없으면 조용히 스킵.
                        if shutil.which('SetFile'):
                            subprocess.run(['SetFile', '-d', mac_date, filepath],
                                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    except Exception:
                        pass
                # macOS xattr: kMDItemWhereFroms (소스 URL 메타데이터)
                post_url_for_xattr = ''
                if post_data and len(post_data) > 1:
                    post_url_for_xattr = str(post_data[1]) if post_data[1] else ''
                if not post_url_for_xattr and uri:
                    # fallback: construct from URI
                    post_url_for_xattr = f"https://bsky.app/profile/{author_handle}/post/{post_id}" if author_handle else ''
                if post_url_for_xattr and sys.platform == 'darwin':
                    try:
                        import plistlib
                        plist_data = plistlib.dumps([post_url_for_xattr, url], fmt=plistlib.FMT_XML)
                        import ctypes, ctypes.util
                        libc = ctypes.CDLL(ctypes.util.find_library('c'))
                        libc.setxattr(
                            filepath.encode('utf-8'),
                            b'com.apple.metadata:kMDItemWhereFroms',
                            plist_data, len(plist_data), 0, 0
                        )
                    except:
                        pass
                # EXIF metadata (for JPG)
                if filepath.lower().endswith(('.jpg', '.jpeg')):
                    try:
                        add_exif(filepath, post_data)
                    except:
                        pass
                # Finder comment (all files including JPG)
                try:
                    add_finder_comment(filepath, post_data)
                except:
                    pass
                # Mirror to _complete folder
                try:
                    complete_path = os.path.join(complete_dir, filename)
                    if not os.path.exists(complete_path):
                        shutil.copy2(filepath, complete_path)
                        # Apply Finder comment + xattr to _complete copy
                        try:
                            add_finder_comment(complete_path, post_data)
                        except:
                            pass
                        if post_url_for_xattr and sys.platform == 'darwin':
                            try:
                                import plistlib
                                plist_data = plistlib.dumps([post_url_for_xattr, url], fmt=plistlib.FMT_XML)
                                import ctypes, ctypes.util
                                libc = ctypes.CDLL(ctypes.util.find_library('c'))
                                libc.setxattr(
                                    complete_path.encode('utf-8'),
                                    b'com.apple.metadata:kMDItemWhereFroms',
                                    plist_data, len(plist_data), 0, 0
                                )
                            except:
                                pass
                except:
                    pass
                return filepath
            except Exception as e:
                log(f"Download failed: {e}")
                return None

        # Images
        images = getattr(embed, 'images', None)
        if images:
            for i, img in enumerate(images):
                url = getattr(img, 'fullsize', '') or getattr(img, 'thumb', '')
                if url:
                    ext = '.jpg'
                    if '.png' in url: ext = '.png'
                    elif '.gif' in url: ext = '.gif'
                    elif '.webp' in url: ext = '.webp'
                    dl(url, make_filename(str(i) if len(images) > 1 else '', ext))

        # Video
        video = getattr(embed, 'video', None)
        if video:
            playlist = getattr(video, 'playlist', '')
            if playlist:
                dl(playlist, make_filename('video', '.mp4'))

        # Record with media
        rmedia = getattr(embed, 'media', None)
        if rmedia:
            rimages = getattr(rmedia, 'images', None)
            if rimages:
                for i, img in enumerate(rimages):
                    url = getattr(img, 'fullsize', '') or getattr(img, 'thumb', '')
                    if url:
                        ext = '.jpg'
                        if '.png' in url: ext = '.png'
                        dl(url, make_filename(f'qm{i}', ext))

        # External thumb
        ext_embed = getattr(embed, 'external', None)
        if ext_embed:
            thumb = getattr(ext_embed, 'thumb', '')
            if thumb:
                dl(thumb, make_filename('thumb', '.jpg'))

        return count

    # EXIF metadata
    def add_exif(filepath, post_data):
        try:
            import piexif
            from PIL import Image
            img = Image.open(filepath)
            exif_dict = piexif.load(img.info.get('exif', b'')) if 'exif' in img.info else {"0th":{}, "Exif":{}, "1st":{}}

            author = post_data[5] if len(post_data) > 5 else ''
            text = (post_data[2] if len(post_data) > 2 else '')[:200]
            url = post_data[1] if len(post_data) > 1 else ''
            date_str = post_data[4] if len(post_data) > 4 else ''

            exif_dict['0th'][piexif.ImageIFD.Artist] = f"@{author}".encode()
            exif_dict['0th'][piexif.ImageIFD.ImageDescription] = text.encode()
            exif_dict['0th'][piexif.ImageIFD.Copyright] = f"Bluesky @{author}".encode()

            if date_str:
                exif_date = date_str.replace('/', ':').replace(' ', ' ')[:19]
                exif_dict['Exif'][piexif.ExifIFD.DateTimeOriginal] = exif_date.encode()
                exif_dict['0th'][piexif.ImageIFD.DateTime] = exif_date.encode()

            if url:
                exif_dict['Exif'][piexif.ExifIFD.UserComment] = piexif.helper.UserComment.dump(url, encoding='ascii')

            exif_bytes = piexif.dump(exif_dict)
            img.save(filepath, quality=95, exif=exif_bytes)
        except ImportError:
            # Fallback: Finder comment
            add_finder_comment(filepath, post_data)
        except:
            pass

    # Finder comment (macOS)
    def add_finder_comment(filepath, post_data):
        try:
            import subprocess
            author = post_data[5] if len(post_data) > 5 else ''
            text = (post_data[2] if len(post_data) > 2 else '')[:100]
            url = post_data[1] if len(post_data) > 1 else ''
            comment = f"@{author}: {text}\\n{url}"
            subprocess.run(['osascript', '-e',
                f'tell application "Finder" to set comment of (POSIX file "{filepath}" as alias) to "{comment}"'],
                capture_output=True, timeout=5)
        except:
            pass

    # Save Excel
    def save_xlsx(filename, data, headers):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Posts"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0085FF", end_color="0085FF", fill_type="solid")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            # Sort by created_at desc
            try:
                data.sort(key=lambda x: x[4] if len(x) > 4 else '', reverse=True)
            except:
                pass

            for row_idx, row_data in enumerate(data, 2):
                for col_idx, val in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)

            wb.save(filename)
            return True
        except Exception as e:
            log(f"Excel save error: {e}")
            # Fallback: TSV
            try:
                with open(filename.replace('.xlsx', '.txt'), 'w') as f:
                    f.write('\t'.join(headers) + '\n')
                    for row in data:
                        f.write('\t'.join(str(v) for v in row) + '\n')
                return True
            except:
                return False

    # ═══ Collection handlers ═══

    def apply_rl_settings(args):
        """커맨드에서 rate limit 설정 적용"""
        mode = args.get("rl_mode", "wait")
        mins = args.get("rl_wait_mins", 15)
        set_rate_limit_mode(mode, mins)

    def collect_posts(args):
        apply_rl_settings(args)
        target = args.get("target", "")
        save_path = args.get("save_path", ".")
        do_media = args.get("download_media", True)
        do_exif = args.get("exif", True)

        short_handle, api_handle = norm_handle(target)
        user_dir = os.path.join(save_path, "bluesky", short_handle)
        media_dir = os.path.join(user_dir, "media")
        excel_dir = os.path.join(user_dir, "excel")
        os.makedirs(media_dir, exist_ok=True)
        os.makedirs(excel_dir, exist_ok=True)

        log(f"Target: @{short_handle}")
        log(f"Save: {user_dir}")

        # Get profile info
        try:
            profile = client.get_profile(actor=api_handle)
            posts_count = getattr(profile, 'posts_count', 0) or 0
            followers_count = getattr(profile, 'followers_count', 0) or 0
            follows_count = getattr(profile, 'follows_count', 0) or 0
            log(f"Posts: {posts_count} | Followers: {followers_count} | Following: {follows_count}")

            # Save profile info
            save_profile(user_dir, profile, short_handle)

            # Download avatar & banner (with archive on change)
            avatar_url = getattr(profile, 'avatar', '')
            banner_url = getattr(profile, 'banner', '')
            import urllib.request, hashlib
            def archive_download(url, path):
                """Download image, archive old version if changed."""
                try:
                    tmp = path + ".tmp"
                    urllib.request.urlretrieve(url, tmp)
                    if os.path.exists(path):
                        # Compare MD5
                        with open(path, 'rb') as f: old_md5 = hashlib.md5(f.read()).hexdigest()
                        with open(tmp, 'rb') as f: new_md5 = hashlib.md5(f.read()).hexdigest()
                        if old_md5 == new_md5:
                            os.remove(tmp)
                            return  # Same file
                        # Archive old version
                        from datetime import datetime
                        date_str = datetime.now().strftime("%Y-%m-%d")
                        archive_dir = os.path.join(os.path.dirname(path), "archive", date_str)
                        os.makedirs(archive_dir, exist_ok=True)
                        os.rename(path, os.path.join(archive_dir, os.path.basename(path)))
                        os.rename(tmp, path)
                        log(f"Profile changed → archived to {date_str}/")
                    else:
                        os.rename(tmp, path)
                except Exception as e:
                    if os.path.exists(tmp): os.remove(tmp)
            # Date-based profile directory
            from datetime import datetime as _dt_main
            _main_pdir = os.path.join(user_dir, "profiles", _dt_main.now().strftime("%Y-%m-%d"))
            os.makedirs(_main_pdir, exist_ok=True)
            if avatar_url:
                _ap = os.path.join(_main_pdir, f"{short_handle}_avatar.jpg")
                if not os.path.exists(_ap):
                    try: urllib.request.urlretrieve(avatar_url, _ap)
                    except: pass
            if banner_url:
                _bp = os.path.join(_main_pdir, f"{short_handle}_banner.jpg")
                if not os.path.exists(_bp):
                    try: urllib.request.urlretrieve(banner_url, _bp)
                    except: pass
        except Exception as e:
            log(f"Profile error: {e}")
            return {"error": str(e)}

        # Collect all posts
        all_data = []
        cursor = None
        total_media = 0
        page = 0

        while True:
            if stop_requested or check_stdin_for_quit():
                log("⛔ 중지 요청 감지 — 포스트 수집 중단")
                break
            try:
                resp = client.get_author_feed(
                    actor=api_handle, limit=100, cursor=cursor,
                    filter='posts_and_author_threads'
                )
            except Exception as e:
                err = str(e).lower()
                if 'rate' in err or '429' in err:
                    # 중간 저장 (AINU 방식)
                    if all_data:
                        _h = ["type", "uri", "post_url", "text", "language", "created_at",
                              "author_handle", "author_name", "likes", "reposts", "replies",
                              "has_media", "media_count", "media_urls"]
                        save_xlsx(os.path.join(excel_dir, f"{short_handle}_posts.xlsx"), all_data, _h)
                        log(f"💾 중간 저장: {len(all_data)}건")
                    if not handle_rate_limit():
                        break
                    continue
                log(f"API error: {e}")
                break

            feed = getattr(resp, 'feed', [])
            if not feed:
                break

            for item in feed:
                post = getattr(item, 'post', None)
                if not post:
                    continue

                # Check if repost
                is_repost = hasattr(item, 'reason') and item.reason is not None

                data = process_post(post)
                if not data:
                    continue

                if is_repost:
                    data.insert(0, 'Repost')
                    # repost_time: when the repost happened (from reason.indexed_at)
                    reason = getattr(item, 'reason', None)
                    repost_time = ''
                    if reason:
                        indexed = getattr(reason, 'indexed_at', '') or getattr(reason, 'created_at', '')
                        if indexed:
                            try:
                                from datetime import timedelta
                                rdt = datetime.fromisoformat(str(indexed).replace('Z', '+00:00'))
                                jst = rdt.astimezone(timezone(timedelta(hours=9)))
                                repost_time = jst.strftime('%Y/%m/%d %H:%M:%S')
                            except:
                                repost_time = str(indexed)
                    data.append(repost_time)
                else:
                    data.insert(0, 'Post')
                    data.append('')  # no repost_time for original posts

                all_data.append(data)

                # Log each post like Twitter does
                ptype = data[0]  # Post or Repost
                pdate = data[5] if len(data) > 5 else ''  # created_at
                ptext = (data[3] if len(data) > 3 else '')[:80]  # text preview
                pauthor = data[6] if len(data) > 6 else ''  # author_handle
                pmedia = data[12] if len(data) > 12 else 0  # media_count
                plikes = data[8] if len(data) > 8 else 0  # likes
                media_tag = f" 📎{pmedia}" if pmedia else ""
                rt_tag = f" ♻️@{pauthor}" if ptype == 'Repost' else ""
                log(f"[{len(all_data)}] {pdate} ❤{plikes}{media_tag}{rt_tag} {ptext}")

                if do_media:
                    new_media = download_media_files(post, media_dir, data[1:], action_type="posts")  # skip type column
                    total_media += new_media

            page += 1
            progress(len(all_data), total_media)

            cursor = getattr(resp, 'cursor', None)
            if not cursor:
                break

            # 아이누 방식 진행 로그 + 중간저장
            n = len(all_data)
            if n > 0 and n % 1000 == 0:
                headers = ["type", "uri", "post_url", "text", "language", "created_at",
                          "author_handle", "author_name", "likes", "reposts", "replies",
                          "has_media", "media_count", "media_urls"]
                save_xlsx(os.path.join(excel_dir, f"{short_handle}_posts.xlsx"), all_data, headers)
                log(f"💾 중간 저장: {n}개")
                time.sleep(5)
            elif n > 0 and n % 500 == 0:
                log(f"📊 진행: {n}개, 미디어 {total_media}개")
                time.sleep(2)
            else:
                time.sleep(1)  # 아이누: 기본 1초

        # Final save
        if all_data:
            headers = ["type", "uri", "post_url", "text", "language", "created_at",
                      "author_handle", "author_name", "likes", "reposts", "replies",
                      "has_media", "media_count", "media_urls", "repost_time"]
            save_xlsx(os.path.join(excel_dir, f"{short_handle}_posts.xlsx"), all_data, headers)
            log(f"Excel saved: {len(all_data)} posts")

        log(f"Complete! Posts: {len(all_data)}, Media: {total_media}")
        progress(len(all_data), total_media, "완료")
        return {"status": "ok", "posts": len(all_data), "media": total_media}

    def collect_likes(args):
        apply_rl_settings(args)
        target = args.get("target", "")
        save_path = args.get("save_path", ".")
        do_media = args.get("download_media", True)

        short_handle, api_handle = norm_handle(target)
        user_dir = os.path.join(save_path, "bluesky", short_handle)
        media_dir = os.path.join(user_dir, "media")
        excel_dir = os.path.join(user_dir, "excel")
        os.makedirs(media_dir, exist_ok=True)
        os.makedirs(excel_dir, exist_ok=True)

        log(f"Collecting likes for @{short_handle}")

        try:
            profile = client.get_profile(actor=api_handle)
            user_did = profile.did
        except Exception as e:
            return {"error": str(e)}

        all_data = []
        cursor = None
        total_media = 0

        while True:
            if stop_requested or check_stdin_for_quit():
                log("⛔ 중지 요청 감지 — 좋아요 수집 중단")
                break
            try:
                params = {'actor': user_did, 'limit': 100}
                if cursor:
                    params['cursor'] = cursor
                resp = client.app.bsky.feed.get_actor_likes(params)
            except Exception as e:
                err = str(e).lower()
                if 'rate' in err or '429' in err:
                    if not handle_rate_limit():
                        break
                    continue
                log(f"API error: {e}")
                break

            feed = getattr(resp, 'feed', [])
            if not feed:
                break

            for item in feed:
                post = getattr(item, 'post', None)
                if not post:
                    continue
                data = process_post(post)
                if data:
                    all_data.append(data)

                    # Log each liked post
                    pdate = data[4] if len(data) > 4 else ''
                    ptext = (data[2] if len(data) > 2 else '')[:80]
                    pauthor = data[5] if len(data) > 5 else ''
                    pmedia = data[11] if len(data) > 11 else 0
                    plikes = data[7] if len(data) > 7 else 0
                    media_tag = f" 📎{pmedia}" if pmedia else ""
                    log(f"[{len(all_data)}] {pdate} @{pauthor} ❤{plikes}{media_tag} {ptext}")

                    if do_media:
                        new_media = download_media_files(post, media_dir, data, action_type="likes")
                        total_media += new_media

            progress(len(all_data), total_media)

            cursor = getattr(resp, 'cursor', None)
            if not cursor:
                break
            time.sleep(2)  # 아이누: 2초 고정

        if all_data:
            headers = ["uri", "post_url", "text", "language", "created_at",
                      "author_handle", "author_name", "likes", "reposts", "replies",
                      "has_media", "media_count", "media_urls", "repost_time"]
            save_xlsx(os.path.join(excel_dir, f"{short_handle}_likes.xlsx"), all_data, headers)
            log(f"Excel saved: {len(all_data)} likes")

        return {"status": "ok", "count": len(all_data), "media": total_media}

    def collect_followers(args):
        apply_rl_settings(args)
        target = args.get("target", "")
        save_path = args.get("save_path", ".")

        short_handle, api_handle = norm_handle(target)
        user_dir = os.path.join(save_path, "bluesky", short_handle)
        profile_dir = os.path.join(user_dir, "profiles")
        excel_dir = os.path.join(user_dir, "excel")
        from datetime import datetime as _dt
        _date_dir = os.path.join(profile_dir, _dt.now().strftime("%Y-%m-%d"))
        os.makedirs(_date_dir, exist_ok=True)
        os.makedirs(excel_dir, exist_ok=True)

        log(f"Collecting followers for @{short_handle}")

        all_data = []
        cursor = None
        import urllib.request

        while True:
            if stop_requested or check_stdin_for_quit():
                log("⛔ 중지 요청 감지 — 팔로워 수집 중단")
                break
            try:
                resp = client.get_followers(actor=api_handle, limit=100, cursor=cursor)
            except Exception as e:
                err = str(e).lower()
                if 'rate' in err or '429' in err:
                    if not handle_rate_limit():
                        break
                    continue
                log(f"API error: {e}")
                break

            followers = getattr(resp, 'followers', [])
            if not followers:
                break

            for f in followers:
                did = getattr(f, 'did', '')
                fhandle = getattr(f, 'handle', '')
                name = getattr(f, 'display_name', '')
                desc = getattr(f, 'description', '') or ''
                fc = getattr(f, 'followers_count', 0) or 0
                fwc = getattr(f, 'follows_count', 0) or 0
                pc = getattr(f, 'posts_count', 0) or 0
                avatar = getattr(f, 'avatar', '')
                banner = getattr(f, 'banner', '')

                all_data.append([did, fhandle, name, desc[:300], fc, fwc, pc, avatar or '', banner or ''])
                log(f"[{len(all_data)}] @{fhandle} ({name}) 팔로워:{fc} 팔로잉:{fwc} 포스트:{pc}")

                # Download profile pic & banner (date-based folder)
                if avatar:
                    p = os.path.join(_date_dir, f"{fhandle}_avatar.jpg")
                    if not os.path.exists(p):
                        try: urllib.request.urlretrieve(avatar, p)
                        except: pass
                if banner:
                    p = os.path.join(_date_dir, f"{fhandle}_banner.jpg")
                    if not os.path.exists(p):
                        try: urllib.request.urlretrieve(banner, p)
                        except: pass

            progress(len(all_data), 0)

            cursor = getattr(resp, 'cursor', None)
            if not cursor:
                break
            time.sleep(2)  # 아이누: 2초 고정

        if all_data:
            headers = ["DID", "Handle", "Name", "Description", "Followers", "Following", "Posts", "Avatar URL", "Banner URL"]
            save_xlsx(os.path.join(excel_dir, f"{short_handle}_followers.xlsx"), all_data, headers)
            log(f"Excel saved: {len(all_data)} followers")

        return {"status": "ok", "count": len(all_data)}

    def collect_following(args):
        apply_rl_settings(args)
        target = args.get("target", "")
        save_path = args.get("save_path", ".")

        short_handle, api_handle = norm_handle(target)
        user_dir = os.path.join(save_path, "bluesky", short_handle)
        profile_dir = os.path.join(user_dir, "profiles")
        excel_dir = os.path.join(user_dir, "excel")
        from datetime import datetime as _dt
        _date_dir = os.path.join(profile_dir, _dt.now().strftime("%Y-%m-%d"))
        os.makedirs(_date_dir, exist_ok=True)
        os.makedirs(excel_dir, exist_ok=True)

        log(f"Collecting following for @{short_handle}")

        all_data = []
        cursor = None
        import urllib.request

        while True:
            if stop_requested or check_stdin_for_quit():
                log("⛔ 중지 요청 감지 — 팔로잉 수집 중단")
                break
            try:
                resp = client.get_follows(actor=api_handle, limit=100, cursor=cursor)
            except Exception as e:
                err = str(e).lower()
                if 'rate' in err or '429' in err:
                    if not handle_rate_limit():
                        break
                    continue
                log(f"API error: {e}")
                break

            follows = getattr(resp, 'follows', [])
            if not follows:
                break

            for f in follows:
                did = getattr(f, 'did', '')
                fhandle = getattr(f, 'handle', '')
                name = getattr(f, 'display_name', '')
                desc = getattr(f, 'description', '') or ''
                fc = getattr(f, 'followers_count', 0) or 0
                fwc = getattr(f, 'follows_count', 0) or 0
                pc = getattr(f, 'posts_count', 0) or 0
                avatar = getattr(f, 'avatar', '')
                banner = getattr(f, 'banner', '')

                all_data.append([did, fhandle, name, desc[:300], fc, fwc, pc, avatar or '', banner or ''])
                log(f"[{len(all_data)}] @{fhandle} ({name}) 팔로워:{fc} 팔로잉:{fwc} 포스트:{pc}")

                if avatar:
                    p = os.path.join(_date_dir, f"{fhandle}_avatar.jpg")
                    if not os.path.exists(p):
                        try: urllib.request.urlretrieve(avatar, p)
                        except: pass
                if banner:
                    p = os.path.join(_date_dir, f"{fhandle}_banner.jpg")
                    if not os.path.exists(p):
                        try: urllib.request.urlretrieve(banner, p)
                        except: pass

            progress(len(all_data), 0)

            cursor = getattr(resp, 'cursor', None)
            if not cursor:
                break
            time.sleep(2)  # 아이누: 2초 고정

        if all_data:
            headers = ["DID", "Handle", "Name", "Description", "Followers", "Following", "Posts", "Avatar URL", "Banner URL"]
            save_xlsx(os.path.join(excel_dir, f"{short_handle}_following.xlsx"), all_data, headers)
            log(f"Excel saved: {len(all_data)} following")

        return {"status": "ok", "count": len(all_data)}

    def collect_blocks(args):
        save_path = args.get("save_path", ".")
        user_dir = os.path.join(save_path, "bluesky", "_blocks")
        excel_dir = os.path.join(user_dir, "excel")
        os.makedirs(user_dir, exist_ok=True)
        os.makedirs(excel_dir, exist_ok=True)

        all_data = []
        cursor = None

        while True:
            if stop_requested or check_stdin_for_quit():
                log("⛔ 중지 요청 감지 — 차단 목록 수집 중단")
                break
            try:
                params = {'limit': 100}
                if cursor: params['cursor'] = cursor
                resp = client.app.bsky.graph.get_blocks(params)
            except Exception as e:
                log(f"API error: {e}")
                break

            blocks = getattr(resp, 'blocks', [])
            if not blocks: break

            for b in blocks:
                all_data.append([
                    getattr(b, 'did', ''),
                    getattr(b, 'handle', ''),
                    getattr(b, 'display_name', ''),
                    (getattr(b, 'description', '') or '')[:300],
                    getattr(b, 'avatar', '') or '',
                ])

            progress(len(all_data), 0)
            cursor = getattr(resp, 'cursor', None)
            if not cursor: break
            time.sleep(2)  # 아이누: 2초 고정

        if all_data:
            save_xlsx(os.path.join(excel_dir, "blocks.xlsx"), all_data,
                     ["DID", "Handle", "Name", "Description", "Avatar"])

        return {"status": "ok", "count": len(all_data)}

    def collect_mutes(args):
        save_path = args.get("save_path", ".")
        user_dir = os.path.join(save_path, "bluesky", "_mutes")
        excel_dir = os.path.join(user_dir, "excel")
        os.makedirs(user_dir, exist_ok=True)
        os.makedirs(excel_dir, exist_ok=True)

        all_data = []
        cursor = None

        while True:
            if stop_requested or check_stdin_for_quit():
                log("⛔ 중지 요청 감지 — 뮤트 목록 수집 중단")
                break
            try:
                params = {'limit': 100}
                if cursor: params['cursor'] = cursor
                resp = client.app.bsky.graph.get_mutes(params)
            except Exception as e:
                log(f"API error: {e}")
                break

            mutes = getattr(resp, 'mutes', [])
            if not mutes: break

            for m in mutes:
                all_data.append([
                    getattr(m, 'did', ''),
                    getattr(m, 'handle', ''),
                    getattr(m, 'display_name', ''),
                    (getattr(m, 'description', '') or '')[:300],
                    getattr(m, 'avatar', '') or '',
                ])

            progress(len(all_data), 0)
            cursor = getattr(resp, 'cursor', None)
            if not cursor: break
            time.sleep(2)  # 아이누: 2초 고정

        if all_data:
            save_xlsx(os.path.join(excel_dir, "mutes.xlsx"), all_data,
                     ["DID", "Handle", "Name", "Description", "Avatar"])

        return {"status": "ok", "count": len(all_data)}

    def collect_search(args):
        apply_rl_settings(args)
        query = args.get("query", "")
        save_path = args.get("save_path", ".")
        do_media = args.get("download_media", True)

        safe_query = re.sub(r'[^\w\s-]', '', query)[:30]
        user_dir = os.path.join(save_path, "bluesky", f"_search_{safe_query}")
        media_dir = os.path.join(user_dir, "media")
        excel_dir = os.path.join(user_dir, "excel")
        os.makedirs(media_dir, exist_ok=True)
        os.makedirs(excel_dir, exist_ok=True)

        log(f"Searching: {query}")
        all_data = []
        cursor = None
        total_media = 0
        max_pages = 100

        for page in range(max_pages):
            if stop_requested or check_stdin_for_quit():
                log("⛔ 중지 요청 감지 — 검색 수집 중단")
                break
            try:
                params = {'q': query, 'limit': 100}
                if cursor: params['cursor'] = cursor
                resp = client.app.bsky.feed.search_posts(params)
            except Exception as e:
                log(f"Search error: {e}")
                break

            posts = getattr(resp, 'posts', [])
            if not posts: break

            for post in posts:
                data = process_post(post)
                if data:
                    all_data.append(data)
                    if do_media:
                        total_media += download_media_files(post, media_dir, data, action_type="search")

            progress(len(all_data), total_media)
            cursor = getattr(resp, 'cursor', None)
            if not cursor: break
            time.sleep(2)  # 아이누: 2초 고정

        if all_data:
            headers = ["uri", "post_url", "text", "language", "created_at",
                      "author_handle", "author_name", "likes", "reposts", "replies",
                      "has_media", "media_count", "media_urls", "repost_time"]
            save_xlsx(os.path.join(excel_dir, f"search_{safe_query}.xlsx"), all_data, headers)

        return {"status": "ok", "count": len(all_data), "media": total_media}

    def collect_notifications(args):
        save_path = args.get("save_path", ".")
        user_dir = os.path.join(save_path, "bluesky", "_notifications")
        excel_dir = os.path.join(user_dir, "excel")
        os.makedirs(user_dir, exist_ok=True)
        os.makedirs(excel_dir, exist_ok=True)

        all_data = []
        cursor = None

        for _ in range(50):
            if stop_requested or check_stdin_for_quit():
                log("⛔ 중지 요청 감지 — 알림 수집 중단")
                break
            try:
                params = {'limit': 100}
                if cursor: params['cursor'] = cursor
                resp = client.app.bsky.notification.list_notifications(params)
            except Exception as e:
                log(f"API error: {e}")
                break

            notifs = getattr(resp, 'notifications', [])
            if not notifs: break

            for n in notifs:
                author = getattr(n, 'author', None)
                record = getattr(n, 'record', None)
                all_data.append([
                    getattr(n, 'uri', ''),
                    getattr(n, 'reason', ''),
                    str(getattr(n, 'is_read', '')),
                    getattr(author, 'handle', '') if author else '',
                    getattr(author, 'display_name', '') if author else '',
                    (getattr(record, 'text', '') if record else '')[:100],
                    getattr(n, 'indexed_at', ''),
                ])

            progress(len(all_data), 0)
            cursor = getattr(resp, 'cursor', None)
            if not cursor: break
            time.sleep(2)  # 아이누: 2초 고정

        if all_data:
            save_xlsx(os.path.join(excel_dir, "notifications.xlsx"), all_data,
                     ["URI", "Reason", "Read", "Author Handle", "Author Name", "Text", "Date"])

        return {"status": "ok", "count": len(all_data)}

    def collect_messages(args):
        save_path = args.get("save_path", ".")
        user_dir = os.path.join(save_path, "bluesky", "_messages")
        excel_dir = os.path.join(user_dir, "excel")
        os.makedirs(user_dir, exist_ok=True)
        os.makedirs(excel_dir, exist_ok=True)

        all_data = []
        try:
            resp = client.chat.bsky.convo.list_convos({'limit': 100})
            convos = getattr(resp, 'convos', [])
            for c in convos:
                members = getattr(c, 'members', [])
                handles = ', '.join(getattr(m, 'handle', '') for m in members)
                names = ', '.join(getattr(m, 'display_name', '') or '' for m in members)
                last_msg = getattr(c, 'last_message', None)
                text = ''
                if last_msg:
                    text = (getattr(last_msg, 'text', '') or '')[:100]
                all_data.append([
                    getattr(c, 'id', ''),
                    handles, names, text,
                    getattr(c, 'updated_at', '') or ''
                ])
        except Exception as e:
            log(f"Messages error: {e}")

        if all_data:
            save_xlsx(os.path.join(excel_dir, "messages.xlsx"), all_data,
                     ["Convo ID", "Members", "Names", "Last Message", "Updated At"])

        return {"status": "ok", "count": len(all_data)}

    def collect_profile(args):
        target = args.get("target", "")
        save_path = args.get("save_path", ".")

        short_handle, api_handle = norm_handle(target)
        user_dir = os.path.join(save_path, "bluesky", short_handle)
        excel_dir = os.path.join(user_dir, "excel")
        os.makedirs(user_dir, exist_ok=True)
        os.makedirs(excel_dir, exist_ok=True)

        try:
            profile = client.get_profile(actor=api_handle)
            save_profile(user_dir, profile, short_handle)

            # Download avatar & banner (with archive on change)
            avatar_url = getattr(profile, 'avatar', '')
            banner_url = getattr(profile, 'banner', '')
            from datetime import datetime as _dt_p
            _pdir = os.path.join(user_dir, "profiles", _dt_p.now().strftime("%Y-%m-%d"))
            os.makedirs(_pdir, exist_ok=True)
            if avatar_url:
                _ap = os.path.join(_pdir, f"{short_handle}_avatar.jpg")
                if not os.path.exists(_ap):
                    try: urllib.request.urlretrieve(avatar_url, _ap)
                    except: pass
            if banner_url:
                _bp = os.path.join(_pdir, f"{short_handle}_banner.jpg")
                if not os.path.exists(_bp):
                    try: urllib.request.urlretrieve(banner_url, _bp)
                    except: pass

            log(f"Profile saved for @{short_handle}")
            return {"status": "ok"}
        except Exception as e:
            return {"error": str(e)}

    def save_profile(user_dir, profile, short_handle):
        try:
            # ★ excel_dir을 user_dir에서 파생 — 호출 컨텍스트의 클로저에 의존하지 않음
            #   (이전: collect_profile/collect_posts 등에 정의된 excel_dir을 참조하려 했는데
            #    save_profile은 동일 스코프 함수라 그 변수가 안 보여서 NameError 발생)
            excel_dir = os.path.join(user_dir, "excel")
            os.makedirs(excel_dir, exist_ok=True)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Profile"
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 60

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0085FF", end_color="0085FF", fill_type="solid")
            for col, h in enumerate(["Item", "Value"], 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            items = [
                ("Handle", getattr(profile, 'handle', '')),
                ("Display Name", getattr(profile, 'display_name', '')),
                ("Description", getattr(profile, 'description', '')),
                ("DID", getattr(profile, 'did', '')),
                ("Followers", getattr(profile, 'followers_count', 0)),
                ("Following", getattr(profile, 'follows_count', 0)),
                ("Posts", getattr(profile, 'posts_count', 0)),
                ("Avatar URL", getattr(profile, 'avatar', '')),
                ("Banner URL", getattr(profile, 'banner', '')),
                ("Created At", getattr(profile, 'created_at', '')),
                ("Indexed At", getattr(profile, 'indexed_at', '')),
            ]
            for i, (key, val) in enumerate(items, 2):
                ws.cell(row=i, column=1, value=key)
                ws.cell(row=i, column=2, value=str(val) if val else '')

            wb.save(os.path.join(excel_dir, f"{short_handle}_profile.xlsx"))
        except Exception as e:
            log(f"Profile save error: {e}")

    def collect_replies(args):
        """Collect all replies the user posted to other people's posts."""
        apply_rl_settings(args)
        target = args.get("target", "")
        save_path = args.get("save_path", ".")
        do_media = args.get("download_media", True)

        short_handle, api_handle = norm_handle(target)
        user_dir = os.path.join(save_path, "bluesky", short_handle)
        media_dir = os.path.join(user_dir, "media")
        excel_dir = os.path.join(user_dir, "excel")
        os.makedirs(media_dir, exist_ok=True)
        os.makedirs(excel_dir, exist_ok=True)

        log(f"Collecting replies by @{short_handle}")

        all_data = []
        cursor = None
        total_media = 0

        while True:
            if stop_requested or check_stdin_for_quit():
                log("⛔ 중지 요청 감지 — 답글 수집 중단")
                break
            try:
                # posts_with_replies: 포스트 + 답글 포함
                resp = client.get_author_feed(
                    actor=api_handle, limit=100, cursor=cursor,
                    filter='posts_with_replies'
                )
            except Exception as e:
                err = str(e).lower()
                if 'rate' in err or '429' in err:
                    if all_data:
                        _h = ["reply_to_url", "uri", "post_url", "text", "language", "created_at",
                              "author_handle", "author_name", "likes", "reposts", "replies",
                              "has_media", "media_count", "media_urls"]
                        save_xlsx(os.path.join(excel_dir, f"{short_handle}_replies.xlsx"), all_data, _h)
                        log(f"💾 중간 저장: {len(all_data)}건")
                    if not handle_rate_limit():
                        break
                    continue
                log(f"API error: {e}")
                break

            feed = getattr(resp, 'feed', [])
            if not feed:
                break

            for item in feed:
                post = getattr(item, 'post', None)
                if not post:
                    continue

                # 답글만 필터링 (reply가 있는 것)
                reply_info = getattr(item, 'reply', None)
                record = getattr(post, 'record', None)
                reply_ref = getattr(record, 'reply', None) if record else None

                if not reply_info and not reply_ref:
                    continue  # 원본 포스트는 스킵, 답글만

                data = process_post(post)
                if not data:
                    continue

                # 답글 대상 포스트 URL 추출
                reply_to_url = ''
                if reply_info:
                    parent = getattr(reply_info, 'parent', None)
                    if parent:
                        p_author = getattr(parent, 'author', None)
                        p_handle = getattr(p_author, 'handle', '') if p_author else ''
                        p_uri = getattr(parent, 'uri', '') or ''
                        p_id = p_uri.split('/')[-1] if p_uri else ''
                        if p_handle and p_id:
                            reply_to_url = f"https://bsky.app/profile/{p_handle}/post/{p_id}"

                row = [reply_to_url] + data
                all_data.append(row)

                # Log
                pdate = data[4] if len(data) > 4 else ''
                ptext = (data[2] if len(data) > 2 else '')[:80]
                plikes = data[7] if len(data) > 7 else 0
                pmedia = data[11] if len(data) > 11 else 0
                media_tag = f" 📎{pmedia}" if pmedia else ""
                log(f"[{len(all_data)}] {pdate} ❤{plikes}{media_tag} ↩️ {ptext}")

                if do_media:
                    new_media = download_media_files(post, media_dir, data, action_type="replies")
                    total_media += new_media

            progress(len(all_data), total_media)

            cursor = getattr(resp, 'cursor', None)
            if not cursor:
                break

            n = len(all_data)
            if n > 0 and n % 1000 == 0:
                headers = ["reply_to_url", "uri", "post_url", "text", "language", "created_at",
                          "author_handle", "author_name", "likes", "reposts", "replies",
                          "has_media", "media_count", "media_urls"]
                save_xlsx(os.path.join(excel_dir, f"{short_handle}_replies.xlsx"), all_data, headers)
                log(f"💾 중간 저장: {n}개")
                time.sleep(5)
            elif n > 0 and n % 500 == 0:
                log(f"📊 진행: {n}개, 미디어 {total_media}개")
                time.sleep(2)
            else:
                time.sleep(1)

        if all_data:
            headers = ["reply_to_url", "uri", "post_url", "text", "language", "created_at",
                      "author_handle", "author_name", "likes", "reposts", "replies",
                      "has_media", "media_count", "media_urls"]
            save_xlsx(os.path.join(excel_dir, f"{short_handle}_replies.xlsx"), all_data, headers)
            log(f"Excel saved: {len(all_data)} replies")

        log(f"Complete! Replies: {len(all_data)}, Media: {total_media}")
        progress(len(all_data), total_media, "완료")
        return {"status": "ok", "count": len(all_data), "media": total_media}

    def collect_comments(args):
        """Collect all replies/comments on a user's posts using getPostThread recursion."""
        apply_rl_settings(args)
        target = args.get("target", "")
        save_path = args.get("save_path", ".")
        do_media = args.get("download_media", True)

        short_handle, api_handle = norm_handle(target)
        user_dir = os.path.join(save_path, "bluesky", short_handle)
        media_dir = os.path.join(user_dir, "media")
        excel_dir = os.path.join(user_dir, "excel")
        os.makedirs(media_dir, exist_ok=True)
        os.makedirs(excel_dir, exist_ok=True)

        log(f"Collecting comments on @{short_handle}'s posts")

        # Helper: recursively extract all replies from a thread
        def extract_replies(thread_node, depth=0):
            replies_data = []
            replies = getattr(thread_node, 'replies', None)
            if not replies:
                return replies_data
            for reply_node in replies:
                post = getattr(reply_node, 'post', None)
                if not post:
                    continue
                data = process_post(post)
                if data:
                    # Add depth info and parent context
                    data.append(depth)
                    replies_data.append((data, post))
                # Recurse into nested replies (unlimited depth)
                replies_data.extend(extract_replies(reply_node, depth + 1))
            return replies_data

        # Step 1: Get all posts from the user
        log("Phase 1: Fetching user's posts...")
        post_uris = []
        cursor = None
        while True:
            if stop_requested or check_stdin_for_quit():
                log("⛔ 중지 요청 감지 — 댓글 수집 중단")
                break
            try:
                resp = client.get_author_feed(
                    actor=api_handle, limit=100, cursor=cursor,
                    filter='posts_no_replies'
                )
            except Exception as e:
                err = str(e).lower()
                if 'rate' in err or '429' in err:
                    if not handle_rate_limit():
                        break
                    continue
                log(f"API error: {e}")
                break

            feed = getattr(resp, 'feed', [])
            if not feed:
                break

            for item in feed:
                post = getattr(item, 'post', None)
                if not post:
                    continue
                uri = getattr(post, 'uri', '')
                reply_count = getattr(post, 'reply_count', 0) or 0
                if reply_count > 0:
                    post_uris.append((uri, reply_count))

            cursor = getattr(resp, 'cursor', None)
            if not cursor:
                break
            time.sleep(2)  # 아이누: 2초 고정

        log(f"Found {len(post_uris)} posts with replies (total reply indicators: {sum(r for _, r in post_uris)})")

        # Step 2: For each post with replies, fetch the full thread
        log("Phase 2: Fetching reply threads...")
        all_comments = []
        total_media = 0
        import urllib.request

        for idx, (uri, expected_replies) in enumerate(post_uris):
            try:
                # depth=1000 to get as many nested replies as possible
                # The API typically returns up to ~1000 replies per thread
                params = {'uri': uri, 'depth': 1000, 'parentHeight': 0}
                resp = client.app.bsky.feed.get_post_thread(params)
            except Exception as e:
                err = str(e).lower()
                if 'rate' in err or '429' in err:
                    log("Rate limited, waiting 60s...")
                    time.sleep(60)
                    # Retry
                    try:
                        params = {'uri': uri, 'depth': 1000, 'parentHeight': 0}
                        resp = client.app.bsky.feed.get_post_thread(params)
                    except:
                        continue
                else:
                    continue

            thread = getattr(resp, 'thread', None)
            if not thread:
                continue

            # Get parent post info for context
            parent_post = getattr(thread, 'post', None)
            parent_url = ''
            if parent_post:
                p_uri = getattr(parent_post, 'uri', '')
                p_author = getattr(parent_post, 'author', None)
                p_handle = getattr(p_author, 'handle', '') if p_author else ''
                p_id = p_uri.split('/')[-1] if p_uri else ''
                parent_url = f"https://bsky.app/profile/{p_handle}/post/{p_id}"

            # Extract all replies recursively
            replies_with_posts = extract_replies(thread)
            for data, post_obj in replies_with_posts:
                depth = data.pop()  # Remove depth from data, add separately
                row = [parent_url, depth] + data
                all_comments.append(row)

                if do_media:
                    new_media = download_media_files(post_obj, media_dir, data, action_type="comments")
                    total_media += new_media

            if (idx + 1) % 10 == 0 or idx == len(post_uris) - 1:
                progress(len(all_comments), total_media, f"댓글 수집 중... ({idx+1}/{len(post_uris)} posts)")
                log(f"Progress: {idx+1}/{len(post_uris)} posts → {len(all_comments)} comments")

            # Rate limit: be gentle
            time.sleep(0.5)

            # Intermediate save every 1000 comments
            if len(all_comments) >= 1000 and len(all_comments) % 1000 < 50:
                headers = ["parent_post_url", "reply_depth",
                          "uri", "post_url", "text", "language", "created_at",
                          "author_handle", "author_name", "likes", "reposts", "replies",
                          "has_media", "media_count", "media_urls"]
                save_xlsx(os.path.join(excel_dir, f"{short_handle}_comments.xlsx"), all_comments, headers)

        # Final save
        if all_comments:
            headers = ["parent_post_url", "reply_depth",
                      "uri", "post_url", "text", "language", "created_at",
                      "author_handle", "author_name", "likes", "reposts", "replies",
                      "has_media", "media_count", "media_urls"]
            save_xlsx(os.path.join(excel_dir, f"{short_handle}_comments.xlsx"), all_comments, headers)
            log(f"Excel saved: {len(all_comments)} comments")

        log(f"Complete! Comments: {len(all_comments)}, Media: {total_media}")
        progress(len(all_comments), total_media, "완료")
        return {"status": "ok", "count": len(all_comments), "media": total_media}

    # ═══ Command loop ═══
    handlers = {
        "posts": collect_posts,
        "likes": collect_likes,
        "replies": collect_replies,
        "media": lambda a: collect_posts({**a, "download_media": True}),
        "followers": collect_followers,
        "following": collect_following,
        "blocks": collect_blocks,
        "mutes": collect_mutes,
        "search": collect_search,
        "notifications": collect_notifications,
        "messages": collect_messages,
        "profile": collect_profile,
        "comments": collect_comments,
    }

    for line in sys.stdin:
        line = line.strip()
        if not line:
            continue

        try:
            args = json.loads(line)
        except json.JSONDecodeError as e:
            print(json.dumps({"error": f"Invalid JSON: {e}"}), flush=True)
            continue

        action = args.get("action", "")

        if action == "quit":
            print(json.dumps({"status": "bye"}), flush=True)
            break

        if action == "ping":
            print(json.dumps({"status": "pong"}), flush=True)
            continue

        handler = handlers.get(action)
        if handler:
            try:
                rate_limit_count = 0  # reset per action
                adaptive_delay = 1.0  # reset adaptive delay per action
                stop_requested = False  # reset stop flag per action
                result = handler(args)
                print(json.dumps(result), flush=True)
            except Exception as e:
                print(json.dumps({"error": str(e)}), flush=True)
        else:
            print(json.dumps({"error": f"Unknown action: {action}"}), flush=True)

if __name__ == "__main__":
    main()
