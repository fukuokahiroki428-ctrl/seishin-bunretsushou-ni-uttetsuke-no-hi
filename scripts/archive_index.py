#!/usr/bin/env python3
"""
산출물 색인기 — 수집해 둔 파일을 훑어 검색 가능한 SQLite 색인을 만든다.

왜 필요한가:
    수집한 파일이 수만~수십만 개다. 폴더를 뒤져서는 "작년에 저장한 그 작가 그림"
    을 찾을 수 없다. 파일마다 흩어져 있는 정보(EXIF 의 작가·제목·원본 URL,
    엑셀 목록, 파일명, 폴더 구조)를 한 곳에 모아 두면 질문으로 찾을 수 있다.

무엇을 읽나:
    이미지  — EXIF (Artist / ImageDescription / UserComment=원본 URL / 촬영일).
              수집할 때 앱이 원본 게시글 정보를 여기 심어 두었다.
    엑셀    — 작가별 작품 목록(ID·제목·작가·유형 등)을 행 단위로.
    JSON/TXT/HTML — 본문 텍스트 일부.
    그 외    — 파일명·경로·크기·날짜만.

특징:
    - 증분: 경로+크기+수정시각이 그대로면 건너뛴다. 몇 번을 돌려도 안전하다.
    - 중단 안전: 배치마다 커밋한다. Ctrl+C 로 끊고 다시 돌리면 이어서 한다.
    - exiftool 은 한 장씩 부르면 수만 장에서 몇 시간이 걸린다. 파일 목록을
      한꺼번에 넘겨(-@) 한 번에 처리한다.
    - 이미지 '내용' 은 읽지 않는다(번들 모델이 글자 전용). 나중에 비전 모델을
      붙일 자리로 vision_desc 칸을 비워 둔다.

사용:
    python3 archive_index.py <산출물폴더> [--db 경로] [--limit N] [--reset]
"""
import argparse, json, os, sqlite3, subprocess, sys, time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    from archive_ask import default_db
except ImportError:      # 단독으로 떼어 쓸 때를 위한 대비
    def default_db() -> Path:
        if sys.platform == 'darwin':
            base = Path.home() / 'Library/Application Support'
        elif os.name == 'nt':
            base = Path(os.environ.get('APPDATA') or (Path.home() / 'AppData/Roaming'))
        else:
            base = Path(os.environ.get('XDG_DATA_HOME') or (Path.home() / '.local/share'))
        # ★ 이름 통일 뒤 경로는 <APPDATA>/Predormition 이다.
        #   예전 자리(Miyo/Predormition)에 이미 색인이 있으면 그것을 그대로 쓴다 —
        #   색인을 다시 돌리게 만들지 않기 위해서다. 앱이 폴더를 옮기면 자연히 새 자리를 본다.
        new = base / 'Predormition/archive_index.db'
        old = base / 'Miyo/Predormition/archive_index.db'
        if not new.exists() and old.exists():
            return old
        return new

IMG_EXT = {'.jpg', '.jpeg', '.png', '.webp', '.gif', '.bmp', '.tiff', '.avif'}
VID_EXT = {'.mp4', '.webm', '.mkv', '.mov', '.m4v', '.avi'}
TXT_EXT = {'.txt', '.json', '.html', '.htm', '.md', '.description', '.vtt', '.srt'}
XLS_EXT = {'.xlsx', '.xlsm'}
SKIP_NAME = {'.DS_Store', 'Thumbs.db'}
# manifest 는 폴더 전체 목록을 담은 요약본이다. 색인하면 어떤 검색어에도 걸려
# 정작 찾으려는 자료를 밀어낸다 → 제외.
SKIP_PREFIX = ('__CHERNOBYL_MANIFEST__', '__PREDORMITION_MANIFEST__', '__TRAD_MANIFEST__')
EXIF_BATCH = 300          # exiftool 한 번에 넘길 파일 수
TEXT_CAP   = 20000        # 파일 하나에서 가져올 본문 최대 길이


# ── 번들 도구 찾기 ─────────────────────────────────────────────────────────
def find_exiftool() -> str | None:
    here = Path(__file__).resolve()
    cands = [
        # 앱 번들 안에서 실행될 때
        here.parent.parent / 'Resources' / 'tools' / 'exiftool' / 'exiftool',
        # 저장소에서 실행될 때
        here.parent.parent / 'mac' / 'chernobyl' / 'resources' / 'tools' / 'exiftool' / 'exiftool',
        Path('/Applications/Predormition.app/Contents/Resources/tools/exiftool/exiftool'),
    ]
    for c in cands:
        if c.is_file() and os.access(c, os.X_OK):
            return str(c)
    from shutil import which
    return which('exiftool')


def platform_of(rel: Path) -> str:
    """경로에서 플랫폼을 추정한다 — 폴더 이름이 곧 출처인 구조를 이용."""
    known = {'twitter', 'x', 'pixiv', 'fanbox', 'bluesky', 'instagram', 'tumblr',
             'youtube', 'niconico', 'discord', 'spinspin', 'asked'}
    for part in rel.parts:
        p = part.lower()
        if p in known:
            return p
    return ''


# ── DB ─────────────────────────────────────────────────────────────────────
SCHEMA_VERSION = 1     # 칸을 늘리거나 바꿀 때 올린다 — 아래 migrate() 가 보고 처리한다

SCHEMA = """
CREATE TABLE IF NOT EXISTS meta (k TEXT PRIMARY KEY, v TEXT);
CREATE TABLE IF NOT EXISTS files (
  path TEXT PRIMARY KEY, name TEXT, ext TEXT, size INTEGER, mtime INTEGER,
  platform TEXT, kind TEXT,
  artist TEXT, title TEXT, description TEXT, source_url TEXT, taken TEXT,
  vision_desc TEXT,            -- 비전 모델을 붙일 때 채울 자리(지금은 비움)
  body TEXT, indexed_at INTEGER
);
CREATE INDEX IF NOT EXISTS idx_files_platform ON files(platform);
CREATE INDEX IF NOT EXISTS idx_files_artist   ON files(artist);
CREATE INDEX IF NOT EXISTS idx_files_mtime    ON files(mtime);
-- ★ 토큰화는 trigram 이어야 한다. 기본 unicode61 은 CJK 를 통째로 한 토큰으로 잡아
--    'vivoさん' 은 찾아도 'vivo' 는 못 찾고, '先生虐待ユメ…' 안의 '先生虐待' 도 못 찾는다.
--    자료 대부분이 일본어라 그대로면 쓸 수 없다. trigram 은 3글자 이상 부분검색이 된다.
--    (2글자 이하 검색어는 trigram 원리상 안 걸리므로 조회 쪽에서 LIKE 로 보완한다.)
CREATE VIRTUAL TABLE IF NOT EXISTS files_fts USING fts5(
  name, artist, title, description, source_url, body, platform,
  content='files', content_rowid='rowid', tokenize='trigram'
);
CREATE TRIGGER IF NOT EXISTS files_ai AFTER INSERT ON files BEGIN
  INSERT INTO files_fts(rowid,name,artist,title,description,source_url,body,platform)
  VALUES (new.rowid,new.name,new.artist,new.title,new.description,new.source_url,new.body,new.platform);
END;
CREATE TRIGGER IF NOT EXISTS files_ad AFTER DELETE ON files BEGIN
  INSERT INTO files_fts(files_fts,rowid,name,artist,title,description,source_url,body,platform)
  VALUES ('delete',old.rowid,old.name,old.artist,old.title,old.description,old.source_url,old.body,old.platform);
END;
CREATE TRIGGER IF NOT EXISTS files_au AFTER UPDATE ON files BEGIN
  INSERT INTO files_fts(files_fts,rowid,name,artist,title,description,source_url,body,platform)
  VALUES ('delete',old.rowid,old.name,old.artist,old.title,old.description,old.source_url,old.body,old.platform);
  INSERT INTO files_fts(rowid,name,artist,title,description,source_url,body,platform)
  VALUES (new.rowid,new.name,new.artist,new.title,new.description,new.source_url,new.body,new.platform);
END;
"""


def open_db(path: Path, reset: bool) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    if reset and path.exists():
        path.unlink()
    db = sqlite3.connect(str(path))
    db.executescript(SCHEMA)
    # ★ 스키마 판올림 — 나중에 칸(예: 비전 모델 결과)이 늘어도 옛 색인이 그냥 깨지지 않게.
    #   판이 낮으면 여기서 ALTER 로 채우고, 감당 못 할 변경이면 다시 만들라고 알린다.
    cur = db.execute("SELECT v FROM meta WHERE k='schema_version'").fetchone()
    have = int(cur[0]) if cur else 0
    if have == 0:
        db.execute("INSERT OR REPLACE INTO meta(k,v) VALUES('schema_version',?)",
                   (str(SCHEMA_VERSION),))
    elif have < SCHEMA_VERSION:
        print(f"  색인 판올림 {have} → {SCHEMA_VERSION}")
        db.execute("INSERT OR REPLACE INTO meta(k,v) VALUES('schema_version',?)",
                   (str(SCHEMA_VERSION),))
    elif have > SCHEMA_VERSION:
        print(f"  ⚠ 색인이 더 새 판({have})입니다. 앱을 업데이트하거나 --reset 으로 다시 만드세요.")
    db.commit()
    # 대량 삽입 — 안전성은 유지하되(WAL) 매 삽입 fsync 는 피한다.
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA synchronous=NORMAL")
    return db


# ── 추출기 ─────────────────────────────────────────────────────────────────
def read_exif_batch(exiftool: str, paths: list[Path]) -> dict[str, dict]:
    """여러 파일의 EXIF 를 한 번에 읽는다. 한 장씩 부르면 수만 장에서 몇 시간 걸린다."""
    if not exiftool or not paths:
        return {}
    listing = "\n".join(str(p) for p in paths)
    try:
        r = subprocess.run(
            [exiftool, '-j', '-charset', 'filename=utf8', '-fast2',
             '-Artist', '-ImageDescription', '-Description', '-UserComment',
             '-XPComment', '-DateTimeOriginal', '-CreateDate', '-@', '-'],
            input=listing, capture_output=True, text=True, timeout=300)
        out = r.stdout.strip()
        if not out:
            return {}
        data = json.loads(out)
    except Exception:
        return {}
    res = {}
    for d in data:
        src = d.get('SourceFile')
        if src:
            res[os.path.realpath(src)] = d
    return res


def read_xlsx(p: Path) -> str:
    """엑셀은 작가별 작품 목록이라 행 자체가 검색 대상이다."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    except Exception:
        return ''
    chunks = []
    total = 0
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if not cells:
                    continue
                line = ' | '.join(cells)
                chunks.append(line)
                total += len(line)
                if total > TEXT_CAP:
                    raise StopIteration
    except StopIteration:
        pass
    except Exception:
        pass
    finally:
        try: wb.close()
        except Exception: pass
    return '\n'.join(chunks)[:TEXT_CAP]


def read_text(p: Path) -> str:
    try:
        raw = p.read_bytes()[:TEXT_CAP * 3]
    except Exception:
        return ''
    try:
        s = raw.decode('utf-8', 'ignore')
    except Exception:
        return ''
    if p.suffix.lower() in {'.html', '.htm'}:
        try:
            from bs4 import BeautifulSoup
            s = BeautifulSoup(s, 'lxml').get_text(' ', strip=True)
        except Exception:
            pass
    return ' '.join(s.split())[:TEXT_CAP]


def kind_of(ext: str) -> str:
    if ext in IMG_EXT: return 'image'
    if ext in VID_EXT: return 'video'
    if ext in XLS_EXT: return 'sheet'
    if ext in TXT_EXT: return 'text'
    return 'other'


# ── 본체 ───────────────────────────────────────────────────────────────────
def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('root', help='산출물 폴더')
    # DB 경로 규칙은 archive_ask.py 한 곳에만 둔다 — 양쪽이 어긋나면
    # 색인은 만들어지는데 질의는 "없습니다" 가 되는, 찾기 어려운 고장이 난다.
    ap.add_argument('--db', default=str(default_db()))
    ap.add_argument('--limit', type=int, default=0, help='이 개수만 처리(시험용)')
    ap.add_argument('--reset', action='store_true', help='색인을 새로 만든다')
    args = ap.parse_args()

    root = Path(args.root).expanduser()
    if not root.is_dir():
        print(f"폴더가 없습니다: {root}"); return 1

    exiftool = find_exiftool()
    db = open_db(Path(args.db), args.reset)
    known = {r[0]: (r[1], r[2]) for r in db.execute("SELECT path,size,mtime FROM files")}

    print(f"산출물 : {root}")
    print(f"색인   : {args.db}")
    print(f"exiftool: {exiftool or '(없음 — 이미지 EXIF 는 건너뜁니다)'}")
    print(f"기존 색인: {len(known):,}건\n")

    todo, skipped, scanned = [], 0, 0
    t0 = time.time()
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if not d.startswith('.') and d != '#recycle']
        for fn in filenames:
            if fn in SKIP_NAME or fn.startswith('._') or fn.startswith(SKIP_PREFIX):
                continue
            p = Path(dirpath) / fn
            try:
                st = p.stat()
            except OSError:
                continue
            scanned += 1
            prev = known.get(str(p))
            if prev and prev[0] == st.st_size and prev[1] == int(st.st_mtime):
                skipped += 1
                continue
            todo.append((p, st))
            if args.limit and len(todo) >= args.limit:
                break
        if args.limit and len(todo) >= args.limit:
            break

    print(f"훑음 {scanned:,}건 · 변경없음 {skipped:,}건 · 새로 색인 {len(todo):,}건\n")
    if not todo:
        print("할 일이 없습니다."); return 0

    done = 0
    for i in range(0, len(todo), EXIF_BATCH):
        batch = todo[i:i + EXIF_BATCH]
        imgs = [p for p, _ in batch if p.suffix.lower() in IMG_EXT]
        exif = read_exif_batch(exiftool, imgs) if imgs else {}

        rows = []
        for p, st in batch:
            ext = p.suffix.lower()
            kind = kind_of(ext)
            rel = p.relative_to(root)
            artist = title = desc = url = taken = ''
            body = ''
            if kind == 'image':
                d = exif.get(os.path.realpath(str(p)), {})
                artist = str(d.get('Artist') or '')
                title  = str(d.get('ImageDescription') or d.get('Description') or '')
                desc   = str(d.get('XPComment') or '')
                uc     = str(d.get('UserComment') or '')
                if uc.startswith('http'):
                    url = uc.split()[0]
                    desc = (desc + ' ' + uc).strip()
                else:
                    desc = (desc + ' ' + uc).strip()
                taken = str(d.get('DateTimeOriginal') or d.get('CreateDate') or '')
            elif kind == 'sheet':
                body = read_xlsx(p)
            elif kind == 'text':
                body = read_text(p)

            rows.append((
                str(p), p.name, ext, st.st_size, int(st.st_mtime),
                platform_of(rel), kind, artist, title, desc.strip(), url, taken,
                None, body, int(time.time())
            ))

        db.executemany("""INSERT INTO files
            (path,name,ext,size,mtime,platform,kind,artist,title,description,
             source_url,taken,vision_desc,body,indexed_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(path) DO UPDATE SET
              size=excluded.size, mtime=excluded.mtime, platform=excluded.platform,
              kind=excluded.kind, artist=excluded.artist, title=excluded.title,
              description=excluded.description, source_url=excluded.source_url,
              taken=excluded.taken, body=excluded.body, indexed_at=excluded.indexed_at
        """, rows)
        db.commit()                      # 배치마다 커밋 — 끊겨도 여기까지는 남는다
        done += len(batch)
        el = time.time() - t0
        rate = done / el if el > 0 else 0
        left = (len(todo) - done) / rate if rate > 0 else 0
        print(f"\r  색인 {done:,}/{len(todo):,}  ({rate:.0f}건/초, 남은 시간 {left/60:.1f}분)",
              end='', flush=True)

    print("\n")
    tot = db.execute("SELECT COUNT(*) FROM files").fetchone()[0]
    print(f"완료 — 색인 총 {tot:,}건")
    for k, c in db.execute("SELECT kind,COUNT(*) FROM files GROUP BY kind ORDER BY 2 DESC"):
        print(f"    {k:6} {c:,}")
    withmeta = db.execute("SELECT COUNT(*) FROM files WHERE artist<>'' OR title<>''").fetchone()[0]
    print(f"  작가/제목 있는 파일: {withmeta:,}건")
    db.close()
    return 0


if __name__ == '__main__':
    sys.exit(main())
